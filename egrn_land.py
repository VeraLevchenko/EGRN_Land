# Программа возвращает правообладателя зу из выписки ЕГРН и вид права
# Выписки хранятся в папке ЕГРН текущего каталога
# В результате создает таблицу в указанной папке и открывает ее.
# Исходная таблица должна быть .xlsx и содержать 2 колонки:
# id	Кадастровыйномер    Видправа    Землепользователь
# ---------------------------------------------------------
# 12    42:30:0501005:113   аренда       Иванов Иван Иванович
#
import os
import random
import openpyxl
import time
import xml.etree.ElementTree as ET
from datetime import date
import datetime

# функция возвращает список .xml файлов из папки, включая подпапки, кроме proto_.xml
def get_file_list(path_Data):
    filelist = []
    for root, dirs, files in os.walk(path_Data):
        for file in files:
            if file.endswith(".xml") and file != "proto_.xml":
                filelist.append(os.path.join(root, file))
    return filelist

def getFilename(cadnum):
    file_egrn = []
    path_Data = os.path.abspath(os.curdir) + "\ЕГРН"
    filelist = get_file_list(path_Data)
    cadnum = cadnum.replace(":", "_")
    for file in filelist:
        if cadnum in file:
            file_egrn.append(file)
    return file_egrn

def getRight_holders(file_name):
    tree = ET.parse(str(file_name[0]))
    list_right_holders = []
    list_right_types = []
    right_holders_individual = tree.findall('right_records/right_record/right_holders/right_holder/individual')
    for right_holder in right_holders_individual:
        #  ---------------------------возвращает ФИО правообладателя----------------------------
        surname = right_holder.find('surname').text
        name = right_holder.find('name').text
        patronymic = right_holder.find('patronymic').text
        list_right_holders.append(surname)
        list_right_holders.append(name)
        list_right_holders.append(patronymic)
    right_holders_public_formation = tree.findall('right_records/right_record/right_holders/right_holder/public_formation/public_formation_type/municipality')
    for right_holder in right_holders_public_formation:
        #  ---------------------------возвращает Публичного правообладателя----------------------------
        name_pub = right_holder.find('name').text
        list_right_holders.append(name_pub)

    right_holders_public_formation = tree.findall('right_records/right_record/right_data/right_type')
    for right_holder in right_holders_public_formation:
        #  ---------------------------возвращает Публичного правообладателя----------------------------
        right_type = right_holder.find('value').text
        list_right_types.append(right_type)
    return list_right_holders, list_right_types

start = time.time() ## точка отсчета времени
name_file = "результат" + str(random.randint(1, 10000)) + ".xlsx"

wb = openpyxl.load_workbook('src.xlsx')
sheet = wb.active

# Создаем объект таблицы Excel и лист
table_output = openpyxl.Workbook()
sheet_table_output = table_output.active

# Создаем заголовки в результирующей таблице
sheet_table_output.append(('id', 'КадастровыйНомерЗУ', "ВидПраваMapinfo", "ЗемлепользовательMapinfo",
                           "ВидПраваЕГРН", "ЗемлепользовательЕГРН", "ДатаИзменения"))
max_row = sheet.max_row
current_date = datetime.datetime.now()

for i in range(2, max_row):
    message = ""
    list_right_holders = []
    list_right_types = []
    id = sheet[i][0].value
    cadnum = str(sheet[i][1].value)
    right_type_mapinfo = str(sheet[i][2].value)
    right_holder_mapinfo = str(sheet[i][3].value)
    print(cadnum)
    file_name = getFilename(cadnum)
    if file_name:
        list_right_holders, list_right_types = getRight_holders(file_name)
    else:
        message = "файл ЕГРН отсутствует"
    list_right_holders = ' '.join(str(list_right_holder) for list_right_holder in list_right_holders) # Преобразуем его в строку, элементы разделяем "; "
    list_right_types = '; '.join(str(list_right_type) for list_right_type in list_right_types)
    print(list_right_holders)
    print(list_right_types)
    # Заполняем строку данными
    object = []
    object.append(id)
    object.append(cadnum)
    object.append(right_type_mapinfo)
    object.append(right_holder_mapinfo)
    if list_right_holders:
        object.append(list_right_types)
    else:
        object.append(message)
    if list_right_holders:
        object.append(list_right_holders)
    else:
        object.append(message)
    object.append(current_date.strftime('%d.%m.%Y'))
    print(object)



    # Добавляем в результирующую таблицу и сохраняем ее
    sheet_table_output.append(object)
    table_output.save(name_file)

end = time.time() - start #  время работы программы
os.startfile(name_file) # Открываем результирующий файл
print("Время работы программы, сек.:", end) # вывод времени
